"""
Business OS tool functions — formatted wrappers over the business DB layer.
Called by execute_tool in brain/tools.py.
"""
from memory.business import (
    add_business, update_business, get_business, list_businesses,
    add_team_member, add_business_goal, complete_goal,
    set_kpi, update_kpi, get_kpis,
    add_contact, log_interaction, update_contact_stage, set_follow_up,
    get_pipeline, get_follow_ups_due, get_interactions,
    log_financial, get_financials, get_nexel_financials, cash_flow_projection,
)


# ── REGISTRY ──────────────────────────────────────────────────────────────────

def tool_add_business(name, type="", stage="idea", description="", revenue_model="", founded="", mrr=0):
    b = add_business(name, type, stage, description, revenue_model, founded, mrr)
    return f"Business added to Nexel: {b['name']} | Stage: {b['stage']} | MRR: ${b['mrr']:,.0f}"


def tool_update_business(name, **fields):
    try:
        b = update_business(name, **fields)
        return f"{b['name']} updated. Stage: {b['stage']} | MRR: ${b['mrr']:,.0f}"
    except ValueError as e:
        return str(e)


def tool_get_business(name):
    try:
        b = get_business(name)
    except ValueError as e:
        return str(e)
    lines = [
        f"── {b['name'].upper()} ──",
        f"Type: {b['type'] or '—'}  |  Stage: {b['stage']}  |  Active: {'yes' if b['active'] else 'no'}",
        f"Founded: {b['founded'] or '—'}",
        f"MRR: ${b['mrr']:,.0f}  |  ARR: ${b['arr']:,.0f}",
        f"Revenue model: {b['revenue_model'] or '—'}",
        f"Description: {b['description'] or '—'}",
    ]
    if b["team"]:
        lines.append("Team: " + ", ".join(f"{m['name']} ({m['role']})" for m in b["team"]))
    if b["goals"]:
        lines.append("Active goals:")
        for g in b["goals"]:
            due = f" — due {g['target_date']}" if g.get("target_date") else ""
            lines.append(f"  • {g['title']}{due}")
    if b["kpis"]:
        lines.append("KPIs:")
        for k in b["kpis"]:
            pct = f" ({k['current'] / k['target'] * 100:.0f}%)" if k.get("target") else ""
            lines.append(f"  • {k['name']}: {k['current']} / {k['target']} {k['unit'] or ''}{pct} [{k['period']}]")
    return "\n".join(lines)


def tool_nexel_overview():
    businesses = list_businesses()
    if not businesses:
        return "No businesses registered under Nexel yet. Add one with add_business."
    fin = get_nexel_financials()
    lines = ["── NEXEL EMPIRE OVERVIEW ──", ""]
    for b in businesses:
        status = "ACTIVE" if b["active"] else "INACTIVE"
        lines.append(f"  {b['name']}  [{b['stage'].upper()}]  MRR ${b['mrr']:,.0f}  {status}")
    lines.append("")
    lines.append(f"Total Revenue (all time): ${fin['total_revenue']:,.2f}")
    lines.append(f"Total Expenses (all time): ${fin['total_expenses']:,.2f}")
    lines.append(f"Total Profit: ${fin['total_profit']:,.2f}")
    return "\n".join(lines)


def tool_add_team_member(business, name, role=""):
    try:
        return add_team_member(business, name, role)
    except ValueError as e:
        return str(e)


def tool_add_goal(business, title, description="", target_date=""):
    try:
        return add_business_goal(business, title, description, target_date)
    except ValueError as e:
        return str(e)


def tool_complete_goal(business, title):
    try:
        return complete_goal(business, title)
    except ValueError as e:
        return str(e)


# ── KPIs ──────────────────────────────────────────────────────────────────────

def tool_set_kpi(business, kpi_name, target, unit="", period="monthly", current=0):
    try:
        return set_kpi(business, kpi_name, target, unit, period, current)
    except ValueError as e:
        return str(e)


def tool_update_kpi(business, kpi_name, current):
    try:
        return update_kpi(business, kpi_name, current)
    except ValueError as e:
        return str(e)


def tool_kpi_report(business):
    try:
        kpis = get_kpis(business)
    except ValueError as e:
        return str(e)
    if not kpis:
        return f"No KPIs set for {business}. Add some with set_kpi."
    lines = [f"── {business.upper()} KPI REPORT ──", ""]
    for k in kpis:
        pct = (k["current"] / k["target"] * 100) if k.get("target") else 0
        bar_filled = int(pct / 10)
        bar = "█" * bar_filled + "░" * (10 - bar_filled)
        status = "✓ ON TRACK" if pct >= 80 else ("⚡ CLOSE" if pct >= 60 else "✗ OFF TARGET")
        unit = k["unit"] or ""
        lines.append(f"{k['name']} [{k['period'].upper()}]")
        lines.append(f"  {k['current']} / {k['target']} {unit}  [{bar}] {pct:.0f}%  {status}")
        lines.append(f"  Last updated: {k['updated_at'][:10] if k.get('updated_at') else '—'}")
        lines.append("")
    return "\n".join(lines).strip()


# ── CRM ───────────────────────────────────────────────────────────────────────

def tool_add_contact(business, name, role="", company="", email="", phone="", notes="", pipeline_stage="lead"):
    try:
        return add_contact(business, name, role, company, email, phone, notes, pipeline_stage)
    except Exception as e:
        return str(e)


def tool_log_interaction(contact_name, interaction_type, notes, business="", date_str=""):
    return log_interaction(contact_name, interaction_type, notes, business, date_str)


def tool_update_pipeline(contact_name, stage):
    return update_contact_stage(contact_name, stage)


def tool_set_follow_up(contact_name, follow_up_date):
    return set_follow_up(contact_name, follow_up_date)


def tool_show_pipeline(business=""):
    contacts = get_pipeline(business)
    if not contacts:
        return "Pipeline is empty." + (f" No contacts for {business}." if business else "")
    stages = ["lead", "contacted", "negotiating", "closed", "churned"]
    stage_icons = {"lead": "◌", "contacted": "◉", "negotiating": "⚡", "closed": "✓", "churned": "✗"}
    lines = [f"── PIPELINE{'  ' + business.upper() if business else ''} ──", ""]
    for stage in stages:
        stage_contacts = [c for c in contacts if c["pipeline_stage"] == stage]
        if not stage_contacts:
            continue
        lines.append(f"{stage_icons.get(stage, '•')} {stage.upper()} ({len(stage_contacts)})")
        for c in stage_contacts:
            last = f"  last contact: {c['last_contact']}" if c.get("last_contact") else ""
            fu = f"  follow-up: {c['follow_up_date']}" if c.get("follow_up_date") else ""
            lines.append(f"   {c['name']} — {c['role'] or '—'} at {c['company'] or '—'}{last}{fu}")
        lines.append("")
    return "\n".join(lines).strip()


def tool_follow_ups_due(days_since=14):
    contacts = get_follow_ups_due(days_since)
    if not contacts:
        return f"No follow-ups due. Everyone has been contacted within {days_since} days."
    lines = [f"── FOLLOW-UPS DUE ──", ""]
    for c in contacts:
        days = c.get("days_since_contact")
        days_str = f"{int(days)} days ago" if days else "never contacted"
        fu = f"  (follow-up was {c['follow_up_date']})" if c.get("follow_up_date") else ""
        lines.append(f"  {c['name']} — {c['role'] or '—'}  |  last contact: {days_str}{fu}")
    return "\n".join(lines)


def tool_contact_history(contact_name, limit=10):
    interactions = get_interactions(contact_name, limit)
    if not interactions:
        return f"No interaction history found for '{contact_name}'."
    lines = [f"── HISTORY: {contact_name.upper()} ──", ""]
    for i in interactions:
        lines.append(f"[{i['date']}] {i['type'].upper()}")
        lines.append(f"  {i['notes']}")
        lines.append("")
    return "\n".join(lines).strip()


# ── FINANCIALS ────────────────────────────────────────────────────────────────

def tool_log_revenue(business, amount, category="", date_str="", notes=""):
    try:
        return log_financial(business, "revenue", float(amount), category, date_str, notes)
    except ValueError as e:
        return str(e)


def tool_log_expense(business, amount, category="", date_str="", notes=""):
    try:
        return log_financial(business, "expense", float(amount), category, date_str, notes)
    except ValueError as e:
        return str(e)


def tool_financial_summary(business, period="month"):
    try:
        f = get_financials(business, period)
    except ValueError as e:
        return str(e)
    lines = [f"── {business.upper()} FINANCIALS ({period.upper()}) ──", ""]
    lines.append(f"Revenue:  ${f['revenue']:>12,.2f}")
    lines.append(f"Expenses: ${f['expenses']:>12,.2f}")
    lines.append(f"Profit:   ${f['profit']:>12,.2f}  (margin: {f['margin']:.1f}%)")
    lines.append("")
    if f["transactions"]:
        lines.append("Recent transactions:")
        for t in f["transactions"][:10]:
            sign = "+" if t["type"] == "revenue" else "-"
            lines.append(f"  {t['date']}  {sign}${t['amount']:,.2f}  {t['category'] or '—'}  {t['notes'] or ''}")
    return "\n".join(lines)


def tool_nexel_financials():
    fin = get_nexel_financials()
    lines = ["── NEXEL P&L ──", ""]
    for b in fin["businesses"]:
        margin = (b["profit"] / b["revenue"] * 100) if b["revenue"] else 0
        lines.append(f"{b['business']:<20}  Rev ${b['revenue']:>10,.0f}  Exp ${b['expenses']:>10,.0f}  P/L ${b['profit']:>10,.0f}  ({margin:.0f}%)")
    lines.append("─" * 70)
    margin = (fin["total_profit"] / fin["total_revenue"] * 100) if fin["total_revenue"] else 0
    lines.append(f"{'TOTAL':<20}  Rev ${fin['total_revenue']:>10,.0f}  Exp ${fin['total_expenses']:>10,.0f}  P/L ${fin['total_profit']:>10,.0f}  ({margin:.0f}%)")
    return "\n".join(lines)


def tool_cash_flow(business):
    try:
        cf = cash_flow_projection(business)
    except ValueError as e:
        return str(e)
    lines = [f"── {business.upper()} CASH FLOW ──", ""]
    lines.append(f"Avg monthly revenue: ${cf['avg_monthly_revenue']:,.2f}")
    lines.append(f"Avg monthly burn:    ${cf['avg_monthly_burn']:,.2f}")
    lines.append(f"Net monthly:         ${cf['net_monthly']:+,.2f}")
    lines.append(f"Current MRR:         ${cf['mrr']:,.2f}")
    lines.append(f"Status: {cf['status'].upper()}")
    return "\n".join(lines)


# ── TAX + ACCOUNTANT EXPORT (4.7) ────────────────────────────────────────────

def tool_tax_estimate(business, tax_rate_pct=30.0):
    try:
        f = get_financials(business, "year")
    except ValueError as e:
        return str(e)
    profit = f["profit"]
    tax = max(0, profit * tax_rate_pct / 100)
    lines = [f"── {business.upper()} TAX ESTIMATE ({tax_rate_pct}% rate) ──", ""]
    lines.append(f"Annual Revenue:  ${f['revenue']:>12,.2f}")
    lines.append(f"Annual Expenses: ${f['expenses']:>12,.2f}")
    lines.append(f"Taxable Profit:  ${profit:>12,.2f}")
    lines.append(f"Estimated Tax:   ${tax:>12,.2f}")
    lines.append(f"After-Tax:       ${profit - tax:>12,.2f}")
    lines.append("")
    lines.append("Note: This is an estimate. Consult an accountant for actual filing.")
    return "\n".join(lines)


def tool_export_for_accountant(business, output_path=""):
    """Export all financial records for a business to Excel."""
    try:
        f = get_financials(business, "all")
    except ValueError as e:
        return str(e)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import os
        from datetime import date

        path = output_path or os.path.expanduser(f"~/Desktop/{business.replace(' ', '_')}_financials_{date.today().isoformat()}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financials"

        hdr_font = Font(name="Courier New", bold=True, color="00FFFF")
        bg = PatternFill("solid", fgColor="080F14")
        hdr_bg = PatternFill("solid", fgColor="0A1520")

        headers = ["Date", "Type", "Category", "Amount", "Notes"]
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = hdr_font
            c.fill = hdr_bg

        for i, t in enumerate(f["transactions"], 2):
            sign = 1 if t["type"] == "revenue" else -1
            ws.cell(row=i, column=1, value=t["date"]).fill = bg
            ws.cell(row=i, column=2, value=t["type"]).fill = bg
            ws.cell(row=i, column=3, value=t["category"] or "").fill = bg
            ws.cell(row=i, column=4, value=sign * t["amount"]).fill = bg
            ws.cell(row=i, column=5, value=t["notes"] or "").fill = bg

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        for row_data in [
            ("Business", business),
            ("Total Revenue", f["revenue"]),
            ("Total Expenses", f["expenses"]),
            ("Net Profit", f["profit"]),
            ("Margin", f"{f['margin']:.1f}%"),
        ]:
            ws2.append(row_data)

        wb.save(path)
        return f"Accountant export saved to: {path}"
    except ImportError:
        return "openpyxl not installed."
    except Exception as e:
        return f"Export error: {e}"


# ── BUSINESS BRIEFING ─────────────────────────────────────────────────────────

def business_briefing() -> str:
    businesses = list_businesses()
    if not businesses:
        return "No businesses registered. Add Addis Market first: add_business('Addis Market', ...)"

    lines = ["── NEXEL BUSINESS BRIEFING ──", ""]

    for b in businesses:
        if not b["active"]:
            continue
        name = b["name"]
        lines.append(f"▸ {name.upper()} [{b['stage'].upper()}]  MRR ${b['mrr']:,.0f}")

        # KPIs
        try:
            kpis = get_kpis(name)
            if kpis:
                for k in kpis:
                    pct = (k["current"] / k["target"] * 100) if k.get("target") else 0
                    flag = "✓" if pct >= 80 else "⚠"
                    lines.append(f"  {flag} {k['name']}: {k['current']} / {k['target']} {k['unit'] or ''} ({pct:.0f}%)")
        except Exception:
            pass

        # Active goals
        try:
            biz = get_business(name)
            if biz["goals"]:
                for g in biz["goals"][:3]:
                    lines.append(f"  → {g['title']}")
        except Exception:
            pass

        lines.append("")

    # Follow-ups due
    due = get_follow_ups_due(7)
    if due:
        lines.append(f"FOLLOW-UPS DUE ({len(due)}):")
        for c in due[:5]:
            days = c.get("days_since_contact")
            lines.append(f"  • {c['name']} — {int(days) if days else '?'} days since last contact")
        lines.append("")

    # Nexel financials snapshot
    try:
        fin = get_nexel_financials()
        if fin["total_revenue"] > 0:
            lines.append(f"NEXEL P/L (all time): Revenue ${fin['total_revenue']:,.0f}  Profit ${fin['total_profit']:,.0f}")
    except Exception:
        pass

    return "\n".join(lines).strip()
